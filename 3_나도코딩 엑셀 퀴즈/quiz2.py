from openpyxl import load_workbook
wb = load_workbook("scores.xlsx", data_only=True)
ws = wb.active

for x in range(2, 12):
    if int(ws.cell(row=x, column=8).value) >= 90:
        ws.cell(row=x, column=9).value = "A"
    elif int(ws.cell(row=x, column=8).value) >=80:
        ws.cell(row=x, column=9).value = "B"
    elif int(ws.cell(row=x, column=8).value) >=70:
        ws.cell(row=x, column=9).value = "C"
    else:
        ws.cell(row=x, column=9).value = "D"
    if int(ws.cell(row=x, column=2).value) < 5:
        ws.cell(row=x, column=9).value = "F"



wb.save("scores2.xlsx")