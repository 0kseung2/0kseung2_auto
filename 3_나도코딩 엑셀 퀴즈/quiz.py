from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.append(["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"])
ws.append([1,10,8,5,14,26,12])
ws.append([2,7,3,7,15,24,18])
ws.append([3,9,5,8,8,12,4])
ws.append([4,7,8,7,17,21,18])
ws.append([5,7,8,7,16,25,15])
ws.append([6,3,5,8,8,17,0])
ws.append([7,4,9,10,16,27,18])
ws.append([8,6,6,6,15,19,17])
ws.append([9,10,10,9,19,30,19])
ws.append([10,9,8,8,20,25,20])

col_d = ws["D"]
for cell in col_d:
    if isinstance(cell.value, int):
        cell.value = 10

ws["H1"] = "총점"
ws["I1"] = "성적"


for i in range(2, 12):
    ws["H{0}".format(i)] = "=SUM(B{0}:G{1})".format(i, i)

# ws["H{}"] = "=SUM(B{0}:G{1})".format(2, 2)



wb.save("scores.xlsx")